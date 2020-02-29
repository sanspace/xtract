from openpyxl import load_workbook
import datetime
import pprint

def get_adjacent_cell_value(ws, cell):
    return ws.cell(cell.row, cell.column + 1).value

def get_date_from_string(string_value):
    return datetime.datetime.strptime(string_value, r'%d.%m.%Y').date()

def get_data(ws, range, data):
    for col in ws.iter_cols(min_row=1, max_col=1, max_row=15):
        for cell in col:
            if isinstance(cell.value, str) is False:
                continue
            text = cell.value.strip()
            if "Your Customer ID" == text:
                data['cutomerid'] = get_adjacent_cell_value(ws, cell)
            if "Name" == text:
                data['customername'] = get_adjacent_cell_value(ws, cell)
            if "Order Date" == text:
                data['orderdate'] = get_date_from_string(
                    get_adjacent_cell_value(ws, cell)
                )
            if "Site Ready Date" == text:
                data['sitereadydate'] = get_date_from_string(
                    get_adjacent_cell_value(ws, cell)
                )
            if "Erection Date" == text:
                data['erectiondate'] = get_date_from_string(
                    get_adjacent_cell_value(ws, cell)
                )
            if "Shutter Erection Date" == text:
                data['shuttererectiondate'] = get_date_from_string(
                    get_adjacent_cell_value(ws, cell)
                )
    # pprint.pprint(data)

def process_sheet(wb, sheetname, data):
    ws = wb[sheetname]
    data['sheetname'] = sheetname
    get_data(ws, range, data)


def process_book(orders, bookpath):
    bookname = bookpath.split('\\')[-1].split('.')[0]
    OMIT_SHEETS = ["Sheet1", "Sheet2", "Sheet3"]
    wb = load_workbook(bookpath)

    for sheetname in wb.sheetnames:
        if sheetname not in OMIT_SHEETS:
            data = {}
            data['bookname'] = bookname
            process_sheet(wb, sheetname, data)
            orders.append(data)

def get_orders():
    orders = []
    process_book(orders, r'.\\input\\7666 Akila Dhanavel Factory order Form.xlsx')
    process_book(orders, r'.\\input\\7903 Ramakrishna Factory Order Form.xlsx')
    pprint.pprint(orders)

get_orders()
